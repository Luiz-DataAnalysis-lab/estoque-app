"""
engine.py — Motor de análise de estoque
Classificação corrigida: 1152=compra Matriz, NF21832=ajuste inventário
"""
import pandas as pd
import numpy as np
import math
from datetime import date, timedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONSTANTES ────────────────────────────────────────────────────────────────
CFOP_VENDA    = [5102, 6102]
CFOP_REMESSA  = [5904, 6904, 5949, 6949, 5910, 6910]
CFOP_RETORNO  = [1904, 2904, 1949, 2949]
CFOP_COMPRA   = [1152]
CFOP_CONSUMO  = [1556]
NF_AJUSTE     = 21832       # NF de ajuste de inventário anual 2025
BASELINE_DATE = "2025-12-30"
Z_MAP         = {'A': 1.65, 'B': 1.28, 'C': 1.00}
FREQ_DIAS     = {'A': 30,   'B': 60,   'C': 90}
FREQ_LABEL    = {'A': 'Mensal', 'B': 'Bimestral', 'C': 'Trimestral'}
MESES_CMM     = ['2025-11','2025-12','2026-01','2026-02','2026-03','2026-04']
N_MESES       = len(MESES_CMM)


# ── LEITURA E LIMPEZA ─────────────────────────────────────────────────────────
def _num(series):
    return pd.to_numeric(
        series.astype(str).str.replace('.','',regex=False).str.replace(',','.',regex=False),
        errors='coerce')

def carregar_inventario(file):
    inv = pd.read_csv(file, sep=None, engine='python', encoding='latin1')
    tp11 = inv[inv['TP']==11].copy()
    tp11 = tp11[tp11['Produto'].notna()].copy()
    for c in ['Quantidade','V.Unitario','Total']: tp11[c] = _num(tp11[c])
    tp11['Produto'] = tp11['Produto'].astype(str).str.strip()

    tp21 = inv[inv['TP']==21].copy()
    tp21 = tp21[tp21['Produto'].notna() &
                ~tp21['Produto'].astype(str).str.match(r'^\d{11,}$')].copy()
    for c in ['Quantidade','V.Unitario','Total']: tp21[c] = _num(tp21[c])
    tp21['Produto'] = tp21['Produto'].astype(str).str.strip()
    return tp11, tp21


def carregar_movimentacoes(file):
    sai = pd.read_csv(file, sep=None, engine='python', encoding='latin1')
    sai['Data_dt'] = pd.to_datetime(sai['Data'], dayfirst=True, errors='coerce')
    sai['Mes']     = sai['Data_dt'].dt.to_period('M')
    for c in ['Quant.','Total']:
        sai[c+'_num'] = _num(sai[c])
    sai['Produto']   = sai['Produto'].astype(str).str.strip()
    sai['is_lely']   = sai['Cliente/Fornecedor'].astype(str).str.upper().str.contains('LELY', na=False)
    sai['is_filial'] = sai['Cliente/Fornecedor'].astype(str).str.upper().str.contains('FILIAL', na=False)
    sai['is_matriz'] = sai['is_lely'] & ~sai['is_filial']

    def classifica(row):
        if row['Doc.Id'] == NF_AJUSTE:                             return 'ajuste_inventario'
        if row['CFOP'] in CFOP_VENDA and not row['is_lely']:       return 'venda_cliente_final'
        if row['CFOP'] in CFOP_VENDA and row['is_lely']:           return 'faturamento_lely'
        if row['CFOP'] in CFOP_REMESSA and row['is_filial']:       return 'remessa_filial'
        if row['CFOP'] in CFOP_REMESSA and not row['is_filial']:   return 'remessa_outros'
        if row['CFOP'] in CFOP_RETORNO:                            return 'retorno_filial'
        if row['CFOP'] in CFOP_COMPRA  and row['is_matriz']:       return 'compra_matriz'
        if row['CFOP'] in CFOP_CONSUMO:                            return 'consumo_operacional'
        return 'outro'

    sai['tipo'] = sai.apply(classifica, axis=1)
    return sai


# ── CMM ───────────────────────────────────────────────────────────────────────
def calcular_cmm(sai):
    sai_c = sai[sai['tipo']=='venda_cliente_final']
    sai_p = sai_c[sai_c['Mes'].astype(str).isin(MESES_CMM)]

    cm = sai_p.groupby(['Produto','Mes'])['Quant._num'].sum().reset_index()
    cmm = sai_p.groupby('Produto').agg(
        Total_Consumo=('Quant._num','sum'), Meses_Ativos=('Mes','nunique'),
        Descricao_Saida=('Descricao','first'), Unidade=('Und.','first')
    ).reset_index()
    cmm['CMM'] = cmm['Total_Consumo'] / cmm['Meses_Ativos']

    todos = pd.MultiIndex.from_product([cmm['Produto'], MESES_CMM], names=['Produto','Mes'])
    cf = cm.copy(); cf['Mes'] = cf['Mes'].astype(str)
    cf = cf.set_index(['Produto','Mes']).reindex(todos, fill_value=0).reset_index()
    desvio = cf.groupby('Produto')['Quant._num'].std().reset_index()
    desvio.columns = ['Produto','Desvio_Pad']
    cmm = cmm.merge(desvio, on='Produto', how='left')
    cmm['Desvio_Pad'] = cmm['Desvio_Pad'].fillna(0)
    return cmm


# ── LEAD TIME REAL ────────────────────────────────────────────────────────────
def calcular_lead_time(sai):
    compras = sai[sai['tipo']=='compra_matriz'].sort_values(['Produto','Data_dt'])
    lt_list = []
    for prod, grp in compras.groupby('Produto'):
        datas = grp['Data_dt'].dropna().sort_values()
        n = len(datas)
        if n >= 2:
            diffs = datas.diff().dropna().dt.days
            diffs = diffs[diffs > 0]
            lt = round(diffs.mean()) if len(diffs) > 0 else 30
        else:
            lt = 30
        lt = min(lt if not np.isnan(lt) else 30, 90)
        lt_list.append({'Produto': prod, 'Lead_Time_Real': lt,
                        'N_Compras': n, 'Ultima_Compra': datas.max()})
    return pd.DataFrame(lt_list)


# ── PARÂMETROS ────────────────────────────────────────────────────────────────
def calcular_parametros(df, custo_pedido=50, taxa_carr=0.25):
    df = df.copy()
    df['LT_dias']  = df['Lead_Time_Real'].fillna(30)
    df['LT_meses'] = df['LT_dias'] / 30
    df['Fator_Z']  = df['ABC'].map(Z_MAP)
    df['ES'] = (df['Fator_Z'] * df['Desvio_Pad'] * np.sqrt(df['LT_meses'])).apply(
        lambda x: math.ceil(x) if pd.notna(x) else 0)
    df['PP'] = (df['CMM'] * df['LT_meses'] + df['ES']).apply(
        lambda x: math.ceil(x) if pd.notna(x) else 0)
    df['D_anual'] = df['CMM'] * 12
    df['H']   = df['V.Unitario'].fillna(0) * taxa_carr
    df['LEC'] = df.apply(
        lambda r: math.ceil(math.sqrt(2*r['D_anual']*custo_pedido/r['H']))
        if r['H']>0 and r['D_anual']>0 else 0, axis=1)
    df['Est_Max']      = df['PP'] + df['LEC']
    df['Estoque_Atual']= df['Quantidade'].fillna(0)
    df['Status'] = df.apply(
        lambda r: 'ACIMA DO MÁXIMO' if r['Estoque_Atual']>r['Est_Max']
        else ('ABAIXO DO PP'  if r['Estoque_Atual']<r['PP'] else 'OK'), axis=1)
    df['Excesso_Un'] = (df['Estoque_Atual'] - df['Est_Max']).clip(lower=0)
    df['Excesso_R$'] = df['Excesso_Un'] * df['V.Unitario'].fillna(0)
    return df


# ── ABC ───────────────────────────────────────────────────────────────────────
def calcular_abc(df):
    df = df.copy()
    df['Valor_Consumo_Anual'] = df['CMM'] * 12 * df['V.Unitario'].fillna(0)
    df = df.sort_values('Valor_Consumo_Anual', ascending=False).reset_index(drop=True)
    tv = df['Valor_Consumo_Anual'].sum()
    df['Pct_Acum'] = df['Valor_Consumo_Anual'].cumsum() / tv if tv > 0 else 0
    df['ABC'] = df['Pct_Acum'].apply(lambda x: 'A' if x<=0.8 else ('B' if x<=0.95 else 'C'))
    return df


# ── REMESSAS ──────────────────────────────────────────────────────────────────
def calcular_remessas(sai):
    rem = sai[sai['tipo']=='remessa_filial']
    ret = sai[sai['tipo']=='retorno_filial']
    fat = sai[sai['tipo']=='faturamento_lely']

    s_env = rem.groupby('Produto').agg(
        Qtd_Enviada=('Quant._num','sum'), Valor_Enviado=('Total_num','sum'),
        Ultima_Remessa=('Data_dt','max'), Descricao=('Descricao','first')).reset_index()
    s_ret = ret.groupby('Produto').agg(
        Qtd_Retornada=('Quant._num','sum'), Valor_Retornado=('Total_num','sum')).reset_index()
    s_fat = fat.groupby('Produto').agg(
        Qtd_Faturada=('Quant._num','sum'), Valor_Faturado=('Total_num','sum')).reset_index()

    r = s_env.merge(s_ret, on='Produto', how='left').merge(s_fat, on='Produto', how='left')
    for c in ['Qtd_Retornada','Qtd_Faturada','Valor_Retornado','Valor_Faturado']:
        r[c] = r[c].fillna(0)
    r['Qtd_Pendente']   = r['Qtd_Enviada']   - r['Qtd_Retornada'] - r['Qtd_Faturada']
    r['Valor_Pendente'] = r['Valor_Enviado']  - r['Valor_Retornado'] - r['Valor_Faturado']
    em_aberto = r[r['Qtd_Pendente'] > 0].sort_values('Valor_Pendente', ascending=False)
    return r, em_aberto


# ── CONCILIAÇÃO PÓS-BASELINE ──────────────────────────────────────────────────
def calcular_conciliacao(sai, inv_clean):
    baseline = pd.Timestamp(BASELINE_DATE)
    sai_pos  = sai[sai['Data_dt'] > baseline].copy()
    tipos    = ['compra_matriz','venda_cliente_final','remessa_filial','retorno_filial','faturamento_lely']
    prods    = set(inv_clean['Produto'].astype(str)) | \
               set(sai_pos[sai_pos['tipo'].isin(tipos)]['Produto'].astype(str))

    rows = []
    for prod in prods:
        ir = inv_clean[inv_clean['Produto'].astype(str)==prod]
        ss = float(ir['Quantidade'].iloc[0]) if len(ir)>0 else 0
        vu = float(ir['V.Unitario'].iloc[0]) if len(ir)>0 else 0
        dc = ir['Descricao'].iloc[0] if len(ir)>0 else ''
        sp = sai_pos[sai_pos['Produto'].astype(str)==prod]
        ent = sp[sp['tipo']=='compra_matriz']['Quant._num'].sum()
        sai_= sp[sp['tipo']=='venda_cliente_final']['Quant._num'].sum()
        fat = sp[sp['tipo']=='faturamento_lely']['Quant._num'].sum()
        rem = sp[sp['tipo']=='remessa_filial']['Quant._num'].sum()
        ret = sp[sp['tipo']=='retorno_filial']['Quant._num'].sum()
        if ent+sai_+fat+rem+ret == 0: continue
        mov = ent - sai_ - fat - rem + ret
        if not dc: dc = sp['Descricao'].iloc[0] if len(sp)>0 else ''
        rows.append({'Produto':prod,'Descricao':dc,'Saldo_Sistema':ss,
                     'Ent_Matriz':ent,'Sai_Clientes':sai_,'Fat_Lely':fat,
                     'Rem_Filial':rem,'Ret_Filial':ret,'Mov_Liquido':mov,
                     'V_Unitario':vu,'Impacto_R$':abs(mov)*vu})
    return pd.DataFrame(rows).sort_values('Impacto_R$', ascending=False).reset_index(drop=True)


# ── ACURÁCIA (mesma fórmula do painel) ───────────────────────────────────────
def calc_acuracia(sistema, contada):
    """Fórmula: 1 - |contada - sistema| / sistema"""
    if pd.isna(sistema) or pd.isna(contada): return None
    if sistema == 0: return 1.0 if contada == 0 else 0.0
    return 1 - abs(contada - sistema) / sistema


def calcular_acuracia_contagens(df_itens, contagens_df):
    """
    df_itens: DataFrame com colunas Produto ou codigo, ABC, Estoque_Atual ou qtdSistema
    contagens_df: DataFrame com colunas codigo, qtd
    Retorna métricas de acurácia por classe e geral.
    """
    if contagens_df is None or len(contagens_df) == 0:
        return {'geral': None, 'A': None, 'B': None, 'C': None, 'D': None,
                'contados': 0, 'total': len(df_itens)}

    # Normalizar colunas do df_itens
    cols = df_itens.columns.tolist()
    prod_col    = 'codigo'      if 'codigo'      in cols else 'Produto'
    estoque_col = 'qtdSistema'  if 'qtdSistema'  in cols else 'Estoque_Atual'
    abc_col     = 'ABC'         if 'ABC'         in cols else 'abc'

    # Garantir que as colunas existem
    cols_needed = [c for c in [prod_col, abc_col, estoque_col] if c in cols]
    itens_ref = df_itens[cols_needed].copy()
    itens_ref = itens_ref.rename(columns={
        prod_col:    'codigo',
        estoque_col: 'qtdSistema',
        abc_col:     'ABC',
    })
    itens_ref['codigo'] = itens_ref['codigo'].astype(str)

    # Normalizar contagens
    cont = contagens_df.copy()
    cont['codigo'] = cont['codigo'].astype(str)
    cont['qtd']    = pd.to_numeric(cont['qtd'], errors='coerce')

    # Merge
    merged = cont.merge(itens_ref, on='codigo', how='inner')

    if len(merged) == 0:
        return {'geral': None, 'A': None, 'B': None, 'C': None, 'D': None,
                'contados': 0, 'total': len(df_itens)}

    # Garantir colunas após merge
    if 'qtdSistema' not in merged.columns:
        return {'geral': None, 'A': None, 'B': None, 'C': None, 'D': None,
                'contados': 0, 'total': len(df_itens)}

    merged['acc'] = merged.apply(
        lambda r: calc_acuracia(
            r['qtdSistema'] if pd.notna(r.get('qtdSistema')) else 0,
            r['qtd']        if pd.notna(r.get('qtd'))        else 0),
        axis=1)
    merged = merged[merged['acc'].notna()]

    result = {
        'geral':    merged['acc'].mean() if len(merged) > 0 else None,
        'contados': len(merged),
        'total':    len(df_itens),
    }
    abc_col_m = 'ABC' if 'ABC' in merged.columns else 'abc'
    for cls in ['A', 'B', 'C', 'D']:
        sub = merged[merged[abc_col_m] == cls] if abc_col_m in merged.columns else pd.DataFrame()
        result[cls] = sub['acc'].mean() if len(sub) > 0 else None
    return result


# ── PIPELINE PRINCIPAL ────────────────────────────────────────────────────────
def rodar_analise(inv_file, sai_file, custo_pedido=50, taxa_carr=0.25):
    inv_clean, inv_tp21 = carregar_inventario(inv_file)
    sai = carregar_movimentacoes(sai_file)

    cmm   = calcular_cmm(sai)
    df_lt = calcular_lead_time(sai)

    df = cmm.merge(inv_clean[['Produto','Descricao','Unidade','Quantidade',
                               'V.Unitario','Total','Tipo Produto','Classificação',
                               'End. Físico']], on='Produto', how='left')
    df = df.merge(df_lt[['Produto','Lead_Time_Real']], on='Produto', how='left')
    df['Descricao']      = df['Descricao'].fillna(df['Descricao_Saida'])
    df['Quantidade']     = df['Quantidade'].fillna(0)
    df['Lead_Time_Real'] = df['Lead_Time_Real'].fillna(30)

    df = calcular_abc(df)
    df = calcular_parametros(df, custo_pedido, taxa_carr)

    inv_prods = set(inv_clean['Produto'].astype(str))
    df_param  = df[df['Produto'].astype(str).isin(inv_prods)].copy()
    df_giro   = df[~df['Produto'].astype(str).isin(inv_prods)].copy()
    sem_saida = inv_clean[~inv_clean['Produto'].isin(set(cmm['Produto']))].sort_values('Total', ascending=False)

    remessas, em_aberto = calcular_remessas(sai)
    conciliacao = calcular_conciliacao(sai, inv_clean)

    # Itens para o app de contagem (com prioridade calculada)
    hoje = date.today()
    itens_contagem = []
    df_p2 = df_param.copy()
    df_p2['ps'] = df_p2['Status'].map({'ABAIXO DO PP':0,'ACIMA DO MÁXIMO':1,'OK':2})
    for _, row in df_p2.sort_values(['ps','ABC','Valor_Consumo_Anual'],
                                     ascending=[True,True,False]).iterrows():
        cls  = row['ABC']
        prox = (hoje + timedelta(days=FREQ_DIAS[cls])).strftime('%Y-%m-%d')
        und  = str(row.get('Unidade_y') or row.get('Unidade_x') or row.get('Unidade',''))
        if und == 'nan': und = 'UN'
        end  = str(row.get('End. Físico','') or '')
        if end == 'nan': end = ''
        prio = ('urgente' if row['Status']=='ABAIXO DO PP'
                else 'excesso' if row['Status']=='ACIMA DO MÁXIMO' else 'normal')
        itens_contagem.append({
            'codigo':       str(row['Produto']),
            'descricao':    str(row['Descricao'])[:100],
            'unidade':      und,
            'endereco':     end,
            'abc':          cls,
            'frequencia':   FREQ_LABEL[cls],
            'proxContagem': prox,
            'qtdSistema':   float(row['Estoque_Atual']),
            'custoUnit':    float(row['V.Unitario']) if pd.notna(row.get('V.Unitario')) else 0,
            'status':       row['Status'],
            'prioridade':   prio,
        })
    # Sem movimento = Classe D
    for _, row in sem_saida.iterrows():
        end = str(row.get('End. Físico','') or '')
        if end == 'nan': end = ''
        itens_contagem.append({
            'codigo':       str(row['Produto']),
            'descricao':    str(row['Descricao'])[:100],
            'unidade':      str(row.get('Unidade','') or 'UN'),
            'endereco':     end,
            'abc':          'D',
            'frequencia':   'Trimestral',
            'proxContagem': (hoje + timedelta(days=90)).strftime('%Y-%m-%d'),
            'qtdSistema':   float(row['Quantidade']) if pd.notna(row.get('Quantidade')) else 0,
            'custoUnit':    float(row['V.Unitario']) if pd.notna(row.get('V.Unitario')) else 0,
            'status':       'Sem Movimento',
            'prioridade':   'normal',
        })

    return {
        'df_param':       df_param,
        'df_giro':        df_giro,
        'sem_saida':      sem_saida,
        'remessas':       remessas,
        'em_aberto':      em_aberto,
        'inv_clean':      inv_clean,
        'inv_tp21':       inv_tp21,
        'conciliacao':    conciliacao,
        'itens_contagem': itens_contagem,
        'meses_cmm':      MESES_CMM,
        'n_meses':        N_MESES,
    }


# ── ESTILOS EXCEL ─────────────────────────────────────────────────────────────
DARK="1A1714"; GREEN="1A7A4A"; GREEN_L="E8F5EE"
ORANGE="B84C00"; ORANGE_L="FDF0E6"; RED="8A1A1A"; RED_L="F8E6E6"
BLUE="1A4F8A"; BLUE_L="E6EEF8"; GREY="F5F5F5"
AMBER="7A4A00"; AMBER_L="FDF5E6"; TEAL="0F6E56"; TEAL_L="E6F5F1"
PURPLE="4A1A7A"; PURPLE_L="F3EEF8"; WHITE="FFFFFF"

def _bd():
    t = Side(style='thin', color='DDDDDD')
    return Border(left=t, right=t, top=t, bottom=t)

def _hdr(ws, row, cols, color, txt=WHITE):
    for c, (label, width) in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font = Font(bold=True, color=txt, size=9, name="Arial")
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _bd()
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[row].height = 26

def _drow(ws, r, vals, fills=None, fmts=None, bolds=None, centers=None, colors=None):
    for c, val in enumerate(vals, 1):
        cell = ws.cell(row=r, column=c, value=val)
        fc = colors[c-1] if colors and colors[c-1] else DARK
        cell.font = Font(name="Arial", size=9, bold=(bolds[c-1] if bolds else False), color=fc)
        cell.border = _bd()
        ha = "center" if (centers and c in centers) else "left"
        cell.alignment = Alignment(vertical="center", horizontal=ha)
        if fills and fills[c-1]: cell.fill = PatternFill("solid", fgColor=fills[c-1])
        if fmts  and fmts[c-1]:  cell.number_format = fmts[c-1]
    ws.row_dimensions[r].height = 16

def _ttl(ws, r, c, txt, size=11, color=DARK):
    cell = ws.cell(r, c, txt)
    cell.font = Font(name="Arial", bold=True, size=size, color=color)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[r].height = 20

def _sub(ws, r, c, txt):
    cell = ws.cell(r, c, txt)
    cell.font = Font(name="Arial", size=8, color="888888")
    ws.row_dimensions[r].height = 14

ST_FILL = {"OK": GREEN_L, "ACIMA DO MÁXIMO": ORANGE_L, "ABAIXO DO PP": RED_L}
ABC_BG  = {"A": GREEN_L, "B": BLUE_L, "C": GREY, "D": "F0F0F0"}
ABC_FG  = {"A": GREEN,   "B": BLUE,   "C": "555555", "D": "999999"}


def gerar_excel(resultado, contagens_df=None):
    """Gera o Excel com 6 abas + acurácia real se contagens_df fornecido."""
    r   = resultado
    df  = r['df_param']
    gs  = r['sem_saida']
    rem = r['remessas']
    emb = r['em_aberto']
    inv = r['inv_clean']
    tp21= r['inv_tp21']
    con = r['conciliacao']
    meses = r['meses_cmm']
    hoje = date.today()

    # Acurácia real das contagens
    acc = calcular_acuracia_contagens(
        df[['Produto','ABC','Estoque_Atual']].rename(columns={'Produto':'codigo'}),
        contagens_df) if contagens_df is not None else None

    wb = Workbook()

    # ── PAINEL ────────────────────────────────────────────────────────────────
    ws1 = wb.active; ws1.title = "📊 PAINEL"
    ws1.sheet_view.showGridLines = False
    _ttl(ws1,1,1,f"PAINEL DE GESTÃO  ·  {hoje.strftime('%d/%m/%Y')}  ·  Base: {meses[0]} a {meses[-1]}",13)
    _sub(ws1,2,1,"NF 21832 = ajuste inventário (excluída)  ·  LT real por item  ·  Custo pedido R$50  ·  Taxa 25%")

    kpis = [
        ("Itens Inventário",    len(inv),              "",   DARK),
        ("Com Consumo Real",    len(df),               "",   BLUE),
        ("Sem Movimento",       len(gs),               "",   ORANGE),
        ("Valor Inventário",    inv['Total'].sum(),    "R$", DARK),
        ("Excesso Estimado",    df['Excesso_R$'].sum(),"R$", RED),
        ("Remessas em Aberto",  emb['Valor_Pendente'].sum(),"R$",AMBER),
    ]
    r2=4; _ttl(ws1,r2,1,"INDICADORES GERAIS",10); r2+=1
    for i,(label,val,prefix,color) in enumerate(kpis,1):
        col=((i-1)%3)*2+1; rr=r2+((i-1)//3)*3
        for rc in [rr,rr+1]:
            for cc in [col,col+1]:
                ws1.cell(rc,cc).border=_bd()
                ws1.cell(rc,cc).fill=PatternFill("solid",fgColor="F7F7F7")
        ws1.merge_cells(start_row=rr,start_column=col,end_row=rr,end_column=col+1)
        ws1.merge_cells(start_row=rr+1,start_column=col,end_row=rr+1,end_column=col+1)
        lc=ws1.cell(rr,col,label); lc.font=Font(name="Arial",size=8,color="888888"); lc.alignment=Alignment(horizontal="center",vertical="center")
        txt=f"R$ {val:,.0f}" if prefix=="R$" else f"{int(val):,}"
        vc=ws1.cell(rr+1,col,txt); vc.font=Font(name="Arial",bold=True,size=13,color=color); vc.alignment=Alignment(horizontal="center",vertical="center")
        ws1.row_dimensions[rr+1].height=26
    r2+=7

    # Acurácia real (se disponível)
    if acc and acc['geral'] is not None:
        _ttl(ws1,r2,1,"ACURÁCIA REAL (CONTAGEM FÍSICA)",10,RED); r2+=1
        _hdr(ws1,r2,[("Indicador",32),("Resultado",16),("Meta",10)],DARK); r2+=1
        acc_rows=[
            (f"Acurácia Geral ({acc['contados']}/{acc['total']} contados)", acc['geral'],    0.95),
            ("Acurácia Classe A",                                            acc.get('A'),   0.98),
            ("Acurácia Classe B",                                            acc.get('B'),   0.95),
            ("Acurácia Classe C",                                            acc.get('C'),   0.95),
        ]
        for lbl,val,meta in acc_rows:
            if val is None: continue
            cor = GREEN if val>=meta else RED
            _drow(ws1,r2,[lbl,f"{val*100:.1f}%",f"{meta*100:.0f}%"],
                  fills=[GREEN_L if val>=meta else RED_L]*3,centers={2,3},
                  colors=[None,cor,GREEN]); r2+=1
        r2+=1

    # Status + ABC
    _ttl(ws1,r2,1,"STATUS DO ESTOQUE",10); r2+=1
    _hdr(ws1,r2,[("Status",22),("Itens",10),("% dos Itens",14),("Ação",44)],DARK); r2+=1
    for st,clr,ac in [("ABAIXO DO PP",RED_L,"Emitir pedido imediatamente"),
                       ("ACIMA DO MÁXIMO",ORANGE_L,"Investigar — compra antecipada?"),
                       ("OK",GREEN_L,"Monitorar conforme freq. ABC")]:
        cnt=df['Status'].value_counts().get(st,0)
        _drow(ws1,r2,[st,cnt,cnt/max(len(df),1),ac],fills=[clr]*4,fmts=[None,"#,##0","0.0%",None],centers={2,3}); r2+=1
    r2+=1

    _ttl(ws1,r2,1,"CURVA ABC",10); r2+=1
    _hdr(ws1,r2,[("Classe",7),("Itens",8),("Cons. Anual (R$)",22),("% Valor",10),("LT Médio (dias)",18),("Freq.",18)],DARK); r2+=1
    tv=df['Valor_Consumo_Anual'].sum()
    for cls in ["A","B","C"]:
        s=df[df['ABC']==cls]
        lt_m=s['LT_dias'].mean() if len(s)>0 else 30
        clr={"A":GREEN,"B":BLUE,"C":"777777"}[cls]
        _drow(ws1,r2,[cls,len(s),s['Valor_Consumo_Anual'].sum(),s['Valor_Consumo_Anual'].sum()/max(tv,1),round(lt_m,0),FREQ_LABEL[cls]],
              fmts=[None,"#,##0","R$ #,##0.00","0.0%","#,##0",None],centers={1,2,4,5,6})
        ws1.cell(r2,1).fill=PatternFill("solid",fgColor=ABC_BG[cls])
        ws1.cell(r2,1).font=Font(name="Arial",bold=True,size=9,color=clr); r2+=1
    for col in range(1,9): ws1.column_dimensions[get_column_letter(col)].width=18
    ws1.column_dimensions['A'].width=30; ws1.column_dimensions['D'].width=46

    # ── PARÂMETROS ────────────────────────────────────────────────────────────
    ws2=wb.create_sheet("⚙️ PARÂMETROS"); ws2.sheet_view.showGridLines=False; ws2.freeze_panes="A3"
    ws2.cell(1,1,"PARÂMETROS DE REPOSIÇÃO · 🔴 LT>45d = risco · 🔵 LT<15d = ES alto").font=Font(name="Arial",bold=True,size=11,color=DARK)
    ws2.row_dimensions[1].height=20
    cols2=[("Cód. Produto",16),("Descrição",44),("Und.",7),("ABC",6),("End. Físico",12),
           ("Estoque Atual",13),("CMM",10),("Desvio Pad.",10),("LT Real (dias)",13),
           ("Est. Segurança",13),("Ponto Pedido",13),("LEC",8),("Est. Máximo",12),
           ("Custo Unit. (R$)",15),("Cons. Anual (R$)",16),("Status",18)]
    _hdr(ws2,2,cols2,DARK)
    ws2.cell(2,9).fill=PatternFill("solid",fgColor=TEAL)
    for i,(_,rd) in enumerate(df.iterrows(),3):
        vu=rd.get('V.Unitario',0) if pd.notna(rd.get('V.Unitario',np.nan)) else 0
        und=str(rd.get('Unidade_y') or rd.get('Unidade_x') or rd.get('Unidade',''))
        lt=int(rd.get('LT_dias',30)); sf=ST_FILL.get(rd['Status'])
        lt_fill=RED_L if lt>45 else (BLUE_L if lt<15 else None)
        vals=[str(rd['Produto']),rd['Descricao'],und,rd['ABC'],str(rd.get('End. Físico','') or ''),
              rd['Estoque_Atual'],round(rd['CMM'],2),round(rd['Desvio_Pad'],2),lt,
              int(rd['ES']),int(rd['PP']),int(rd['LEC']),int(rd['Est_Max']),vu,rd['Valor_Consumo_Anual'],rd['Status']]
        fmts=[None,None,None,None,None,"#,##0.00","#,##0.00","#,##0.00","#,##0","#,##0","#,##0","#,##0","#,##0","R$ #,##0.00","R$ #,##0.00",None]
        fills=[None,None,None,None,None,None,None,None,lt_fill]+[None]*6+[sf]
        _drow(ws2,i,vals,fills=fills,fmts=fmts,centers={3,4,6,7,8,9,10,11,12,13})
        ws2.cell(i,4).fill=PatternFill("solid",fgColor=ABC_BG.get(rd['ABC'],WHITE))
        ws2.cell(i,4).font=Font(name="Arial",bold=True,size=9,color=ABC_FG.get(rd['ABC'],DARK))
    ws2.auto_filter.ref=f"A2:{get_column_letter(len(cols2))}2"

    # ── AÇÕES ─────────────────────────────────────────────────────────────────
    ws3=wb.create_sheet("🚨 AÇÕES"); ws3.sheet_view.showGridLines=False; ws3.freeze_panes="A4"
    ws3.cell(1,1,"AÇÕES PRIORITÁRIAS — Execute em ordem: 🔴 Comprar → ⚠️ Investigar → 🔄 Cobrar remessa").font=Font(name="Arial",bold=True,size=11,color=DARK)
    ws3.row_dimensions[1].height=20
    cols3=[("Tipo",12),("Ação",20),("Cód. Produto",16),("Descrição",44),("ABC",6),
           ("LT",9),("Estoque Atual",13),("Referência",14),("Qtd",10),("Impacto (R$)",14),("Obs.",30)]
    _hdr(ws3,3,cols3,DARK)
    rn=4
    rup=df[df['Status']=='ABAIXO DO PP'].copy()
    rup['deficit']=(rup['PP']-rup['Estoque_Atual']).clip(lower=0)
    rup['impacto']=rup['deficit']*rup['V.Unitario'].fillna(0)
    for _,rd in rup.sort_values(['ABC','impacto'],ascending=[True,False]).iterrows():
        lt=int(rd.get('LT_dias',30)); bg=RED_L if rd['ABC']=='A' else "FFF5F5"
        obs="⚠️ LT>45d — risco elevado" if lt>45 else ""
        _drow(ws3,rn,["🔴 COMPRAR","Emitir pedido",str(rd['Produto']),rd['Descricao'],
                       rd['ABC'],lt,rd['Estoque_Atual'],int(rd['PP']),int(rd['deficit']),rd['impacto'],obs],
              fills=[bg]*10+[None],fmts=[None,None,None,None,None,"#,##0","#,##0.00","#,##0","#,##0","R$ #,##0.00",None],
              centers={5,6,7,8,9},bolds=[True,True]+[False]*9,colors=[RED,RED]+[None]*9); rn+=1
    exc=df[df['Status']=='ACIMA DO MÁXIMO'].sort_values('Excesso_R$',ascending=False)
    for _,rd in exc.iterrows():
        bg=ORANGE_L if rd['Excesso_R$']>5000 else "FFFBF5"
        _drow(ws3,rn,["⚠️ INVESTIGAR","Revisar necessidade",str(rd['Produto']),rd['Descricao'],
                       rd['ABC'],int(rd.get('LT_dias',30)),rd['Estoque_Atual'],int(rd['Est_Max']),
                       int(rd['Excesso_Un']),rd['Excesso_R$'],"Capital acima do máximo"],
              fills=[bg]*10+[None],fmts=[None,None,None,None,None,"#,##0","#,##0.00","#,##0","#,##0","R$ #,##0.00",None],
              centers={5,6,7,8,9},bolds=[True,True]+[False]*9,colors=[ORANGE,ORANGE]+[None]*9); rn+=1
    for _,rd in emb[emb['Valor_Pendente']>1000].head(30).iterrows():
        ult=rd['Ultima_Remessa'].strftime('%d/%m/%Y') if pd.notna(rd['Ultima_Remessa']) else ''
        _drow(ws3,rn,["🔄 COBRAR","Cobrar retorno Filial",str(rd['Produto']),rd.get('Descricao',''),
                       "—","—","—","—",int(rd['Qtd_Pendente']),rd['Valor_Pendente'],f"Última: {ult}"],
              fills=[AMBER_L]*10+[None],fmts=[None,None,None,None,None,None,None,None,"#,##0","R$ #,##0.00",None],
              centers={5,6,7,8,9},bolds=[True,True]+[False]*9,colors=[AMBER,AMBER]+[None]*9); rn+=1
    ws3.auto_filter.ref=f"A3:{get_column_letter(len(cols3))}3"

    # ── CONCILIAÇÃO ───────────────────────────────────────────────────────────
    ws4=wb.create_sheet("🔄 CONCILIAÇÃO"); ws4.sheet_view.showGridLines=False; ws4.freeze_panes="A4"
    ws4.cell(1,1,f"CONCILIAÇÃO PÓS-BASELINE ({BASELINE_DATE}) — Movimentos por CFOP após ajuste NF 21832").font=Font(name="Arial",bold=True,size=11,color=DARK)
    ws4.row_dimensions[1].height=20
    cols4=[("Cód. Produto",16),("Descrição",44),("Saldo Sistema",13),
           ("+ Entradas\n1152",14),("- Saídas\n5102",14),("- Fat.LELY\n5102",13),
           ("- Remessas\n5904",14),("+ Retornos\n1904",14),("= Mov. Líquido",14),
           ("Custo Unit.",14),("Impacto (R$)",14)]
    _hdr(ws4,3,cols4,DARK)
    for ci,clr in [(4,GREEN),(5,RED),(6,RED),(7,AMBER),(8,TEAL),(9,BLUE)]:
        ws4.cell(3,ci).fill=PatternFill("solid",fgColor=clr)
    for i,(_,rd) in enumerate(con.iterrows(),4):
        mov=rd['Mov_Liquido']
        bg=GREEN_L if abs(mov)<0.01 else (ORANGE_L if abs(mov)<5 else RED_L if mov<-5 else AMBER_L)
        _drow(ws4,i,[str(rd['Produto']),rd['Descricao'],rd['Saldo_Sistema'],
                     rd['Ent_Matriz'],rd['Sai_Clientes'],rd['Fat_Lely'],
                     rd['Rem_Filial'],rd['Ret_Filial'],mov,rd['V_Unitario'],rd['Impacto_R$']],
              fills=[None,None]+[None]*6+[bg,None,None],
              fmts=[None,None,"#,##0.00","#,##0.00","#,##0.00","#,##0.00","#,##0.00","#,##0.00",
                    "#,##0.00;[Red]-#,##0.00","R$ #,##0.00","R$ #,##0.00"],centers={3,4,5,6,7,8,9})
        if mov<-5: ws4.cell(i,9).font=Font(name="Arial",size=9,bold=True,color=RED)
        elif mov>5:ws4.cell(i,9).font=Font(name="Arial",size=9,bold=True,color=GREEN)
    ws4.auto_filter.ref=f"A3:{get_column_letter(len(cols4))}3"

    # ── CONTAGEM ──────────────────────────────────────────────────────────────
    ws5=wb.create_sheet("📋 CONTAGEM"); ws5.sheet_view.showGridLines=False; ws5.freeze_panes="A3"
    ws5.cell(1,1,"PLANO DE CONTAGEM · Preencha col. J · Divergência e Acurácia calculam automaticamente · Oculte col. I antes de imprimir").font=Font(name="Arial",bold=True,size=10,color=DARK)
    ws5.row_dimensions[1].height=18
    cols5=[("Prioridade",12),("Cód. Produto",16),("Descrição",44),("Und.",7),
           ("End. Físico",13),("ABC",6),("Freq.",13),("Próx. Contagem",14),
           ("Qtd. Sistema",12),("Qtd. Contada ✏️",14),
           ("Divergência",12),("Acurácia %",11),("Status",16),("Observação",22)]
    _hdr(ws5,2,cols5,DARK)
    for ci,clr in [(9,PURPLE),(10,GREEN),(11,BLUE),(12,BLUE),(13,BLUE)]:
        ws5.cell(2,ci).fill=PatternFill("solid",fgColor=clr)

    # Preencher contagens existentes se disponíveis
    cont_dict = {}
    if contagens_df is not None and len(contagens_df)>0:
        for _,row in contagens_df.iterrows():
            cont_dict[str(row['codigo'])] = float(row['qtd'])

    df_p3=df.copy(); df_p3['ps']=df_p3['Status'].map({'ABAIXO DO PP':0,'ACIMA DO MÁXIMO':1,'OK':2})
    rn=3
    for _,rd in df_p3.sort_values(['ps','ABC','Valor_Consumo_Anual'],ascending=[True,True,False]).iterrows():
        cls=rd['ABC']
        prio="🔴 URGENTE" if rd['Status']=='ABAIXO DO PP' else ("⚠️ EXCESSO" if rd['Status']=='ACIMA DO MÁXIMO' else f"Classe {cls}")
        prox=(hoje+timedelta(days=FREQ_DIAS[cls])).strftime('%d/%m/%Y')
        und=str(rd.get('Unidade_y') or rd.get('Unidade_x') or '')
        end=str(rd.get('End. Físico','') or '')
        bg={"ABAIXO DO PP":RED_L,"ACIMA DO MÁXIMO":ORANGE_L,"OK":None}.get(rd['Status'])
        qtd_cont=cont_dict.get(str(rd['Produto']),'')
        vals=[prio,str(rd['Produto']),rd['Descricao'],und,end,cls,FREQ_LABEL[cls],prox,rd['Estoque_Atual'],qtd_cont,"","","",""]
        fmts=[None,None,None,None,None,None,None,"DD/MM/YYYY","#,##0.00",None,None,"0.0%",None,None]
        fills=[bg]*8+[PURPLE_L,GREEN_L,BLUE_L,BLUE_L,BLUE_L,None]
        _drow(ws5,rn,vals,fills=fills,fmts=fmts,centers={1,4,6,7,8,9,10,11,12,13})
        for ci,formula,fmt in [
            (11,f'=IF(J{rn}="","",J{rn}-I{rn})','#,##0.00;[Red]-#,##0.00'),
            (12,f'=IF(J{rn}="","",IF(I{rn}=0,IF(J{rn}=0,1,0),1-ABS(J{rn}-I{rn})/I{rn}))','0.0%'),
            (13,f'=IF(J{rn}="","—",IF(ABS(J{rn}-I{rn})<=0.01,"✅ OK",IF(ABS(J{rn}-I{rn})/MAX(I{rn},0.01)<=0.05,"⚠️ Dif. Pequena","❌ Divergência")))',''),
        ]:
            c=ws5.cell(rn,ci); c.value=formula; c.number_format=fmt
            c.font=Font(name="Arial",size=9); c.border=_bd()
            c.alignment=Alignment(horizontal="center",vertical="center")
            c.fill=PatternFill("solid",fgColor=BLUE_L)
        rn+=1
    for _,rd in gs.iterrows():
        qtd_cont=cont_dict.get(str(rd['Produto']),'')
        vals=["Classe D",str(rd['Produto']),rd['Descricao'],str(rd.get('Unidade','') or ''),
              str(rd.get('End. Físico','') or ''),"D","Trimestral",
              (hoje+timedelta(days=90)).strftime('%d/%m/%Y'),rd['Quantidade'],qtd_cont,"","","",""]
        fmts=[None,None,None,None,None,None,None,"DD/MM/YYYY","#,##0.00",None,None,"0.0%",None,None]
        fills=["F5F5F5"]*8+[PURPLE_L,GREEN_L,BLUE_L,BLUE_L,BLUE_L,None]
        _drow(ws5,rn,vals,fills=fills,fmts=fmts,centers={1,4,6,7,8,9,10,11,12,13})
        for ci,formula,fmt in [
            (11,f'=IF(J{rn}="","",J{rn}-I{rn})','#,##0.00;[Red]-#,##0.00'),
            (12,f'=IF(J{rn}="","",IF(I{rn}=0,IF(J{rn}=0,1,0),1-ABS(J{rn}-I{rn})/I{rn}))','0.0%'),
            (13,f'=IF(J{rn}="","—",IF(ABS(J{rn}-I{rn})<=0.01,"✅ OK",IF(ABS(J{rn}-I{rn})/MAX(I{rn},0.01)<=0.05,"⚠️ Dif. Pequena","❌ Divergência")))',''),
        ]:
            c=ws5.cell(rn,ci); c.value=formula; c.number_format=fmt
            c.font=Font(name="Arial",size=9); c.border=_bd()
            c.alignment=Alignment(horizontal="center",vertical="center")
            c.fill=PatternFill("solid",fgColor=BLUE_L)
        rn+=1
    ws5.auto_filter.ref=f"A2:{get_column_letter(len(cols5))}2"

    # ── SEM MOVIMENTO ─────────────────────────────────────────────────────────
    ws6=wb.create_sheet("📦 SEM MOVIMENTO"); ws6.sheet_view.showGridLines=False; ws6.freeze_panes="A3"
    ws6.cell(1,1,f"ITENS SEM CONSUMO REAL — {len(gs)} itens · R$ {gs['Total'].sum():,.2f}").font=Font(name="Arial",bold=True,size=11,color=ORANGE)
    ws6.row_dimensions[1].height=20
    cols6=[("Cód. Produto",16),("Descrição",44),("Tipo Produto",16),("Classificação",22),
           ("Und.",7),("Estoque Atual",13),("Custo Unit. (R$)",15),("Valor Total (R$)",15),("Ação Sugerida",24)]
    _hdr(ws6,2,cols6,ORANGE)
    for i,(_,rd) in enumerate(gs.iterrows(),3):
        val=rd['Total'] if pd.notna(rd['Total']) else 0
        bg=ORANGE_L if val>5000 else ("FFFBF5" if val>1000 else None)
        acao="🔴 Revisar urgente" if val>5000 else ("⚠️ Avaliar descarte" if val>1000 else "📋 Verificar na contagem")
        _drow(ws6,i,[str(rd['Produto']),rd['Descricao'],rd.get('Tipo Produto',''),rd.get('Classificação',''),
                     str(rd.get('Unidade','') or ''),rd['Quantidade'],rd['V.Unitario'],val,acao],
              fills=[None,bg]+[None]*7,fmts=[None,None,None,None,None,"#,##0.00","R$ #,##0.00","R$ #,##0.00",None],centers={5,6})
    ws6.auto_filter.ref=f"A2:{get_column_letter(len(cols6))}2"

    out = BytesIO(); wb.save(out); out.seek(0)
    return out
